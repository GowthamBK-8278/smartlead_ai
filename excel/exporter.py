# -----------------------------------------------------------------------------
# SmartLead AI – Excel Report Exporter
# Styled Excel exports with color-coded HOT/WARM/COLD rows, bold headers,
# auto column widths, and summary stats worksheet.
# -----------------------------------------------------------------------------

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ── Color palette ────────────────────────────────────────────────────────────
_COLORS = {
    "HOT LEAD":  {"row": "FFEBEE", "badge": "C62828"},   # warm red
    "WARM LEAD": {"row": "FFF8E1", "badge": "F57F17"},   # amber
    "COLD LEAD": {"row": "E3F2FD", "badge": "1565C0"},   # blue
}
_HEADER_FILL   = PatternFill("solid", fgColor="1A237E")  # deep navy
_HEADER_FONT   = Font(bold=True, color="FFFFFF", size=11)
_THIN_BORDER   = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _make_reports_dir() -> str:
    base_dir = os.path.join(
        os.path.dirname(os.path.dirname(__file__)), "data", "reports"
    )
    os.makedirs(base_dir, exist_ok=True)
    return base_dir


def export_to_excel(df: pd.DataFrame, domain: str) -> str:
    """
    Export leads DataFrame to a styled Excel file.
    Returns the absolute path to the created file.
    """
    reports_dir = _make_reports_dir()
    file_path = os.path.join(reports_dir, f"{domain}_Leads.xlsx")

    wb = Workbook()

    # ── Sheet 1: Lead Data ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Lead Data"

    export_cols = ["Username", "Platform", "Title", "Comment",
                   "Interest Score", "Lead Category", "Sentiment", "Date"]

    # Filter only columns that exist in df
    available = [c for c in export_cols if c in df.columns]
    ws.append(available)

    # Style header row
    for cell in ws[1]:
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _THIN_BORDER

    ws.row_dimensions[1].height = 22

    # Write data rows
    for _, row in df.iterrows():
        row_data = [row.get(c, "") for c in available]
        ws.append(row_data)

        # Determine row color from Lead Category
        category = str(row.get("Lead Category", "")).upper()
        match = next(
            (v for k, v in _COLORS.items() if k in category), None
        )
        if match:
            row_fill = PatternFill("solid", fgColor=match["row"])
            last_row = ws.max_row
            for col_idx in range(1, len(available) + 1):
                cell = ws.cell(row=last_row, column=col_idx)
                cell.fill      = row_fill
                cell.border    = _THIN_BORDER
                cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Auto column widths
    for col_idx, col_name in enumerate(available, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row,
            min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.title = "Summary"
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15

    summary_data = [
        ("Domain",        domain),
        ("Total Leads",   len(df)),
        ("HOT Leads",     len(df[df.get("Lead Category", pd.Series()) == "HOT LEAD"]) if "Lead Category" in df else 0),
        ("WARM Leads",    len(df[df.get("Lead Category", pd.Series()) == "WARM LEAD"]) if "Lead Category" in df else 0),
        ("COLD Leads",    len(df[df.get("Lead Category", pd.Series()) == "COLD LEAD"]) if "Lead Category" in df else 0),
        ("Platforms",     df["Platform"].nunique() if "Platform" in df else 0),
        ("Avg Score",     round(df["Interest Score"].mean(), 1) if "Interest Score" in df else 0),
        ("Max Score",     df["Interest Score"].max() if "Interest Score" in df else 0),
    ]

    ws2.append(["Metric", "Value"])
    for cell in ws2[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for label, value in summary_data:
        ws2.append([label, value])

    wb.save(file_path)
    return file_path


def export_real_leads(df: pd.DataFrame, domain: str) -> str:
    """
    Export real scraped leads DataFrame to a styled Excel file.
    Returns the absolute path to the created file.
    """
    reports_dir = _make_reports_dir()
    file_path = os.path.join(reports_dir, f"Real_Leads_{domain}.xlsx")

    wb = Workbook()

    # ── Sheet 1: Lead Data ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Scraped Leads"

    export_cols = ["Username", "Platform", "Source Account", "Comment",
                   "Score", "Profile URL", "DM Sent", "Date"]

    # Filter only columns that exist in df
    available = [c for c in export_cols if c in df.columns]
    ws.append(available)

    # Style header row
    for cell in ws[1]:
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _THIN_BORDER

    ws.row_dimensions[1].height = 22

    # Write data rows
    for _, row in df.iterrows():
        row_data = [row.get(c, "") for c in available]
        ws.append(row_data)

        # Determine row color from Score
        score = int(row.get("Score", 0))
        if score >= 80:
            match = _COLORS["HOT LEAD"]
        elif score >= 50:
            match = _COLORS["WARM LEAD"]
        else:
            match = _COLORS["COLD LEAD"]

        row_fill = PatternFill("solid", fgColor=match["row"])
        last_row = ws.max_row
        for col_idx in range(1, len(available) + 1):
            cell = ws.cell(row=last_row, column=col_idx)
            cell.fill      = row_fill
            cell.border    = _THIN_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    # Auto column widths
    for col_idx, col_name in enumerate(available, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(col_name)
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row,
            min_col=col_idx, max_col=col_idx
        ):
            for cell in row:
                try:
                    max_len = max(max_len, len(str(cell.value or "")))
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 60)

    ws.freeze_panes = "A2"

    # ── Sheet 2: Summary ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2.title = "Summary"
    ws2.column_dimensions["A"].width = 25
    ws2.column_dimensions["B"].width = 15

    total_leads = len(df)
    hot_leads = len(df[df["Score"] >= 80]) if "Score" in df.columns else 0
    warm_leads = len(df[(df["Score"] >= 50) & (df["Score"] < 80)]) if "Score" in df.columns else 0
    cold_leads = len(df[df["Score"] < 50]) if "Score" in df.columns else 0
    
    summary_data = [
        ("Domain",        domain),
        ("Total Leads",   total_leads),
        ("HOT Leads",     hot_leads),
        ("WARM Leads",    warm_leads),
        ("COLD Leads",    cold_leads),
        ("Platforms",     df["Platform"].nunique() if "Platform" in df.columns else 0),
        ("Avg Score",     round(df["Score"].mean(), 1) if "Score" in df.columns else 0),
        ("Max Score",     df["Score"].max() if "Score" in df.columns else 0),
    ]

    ws2.append(["Metric", "Value"])
    for cell in ws2[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for label, value in summary_data:
        ws2.append([label, value])

    wb.save(file_path)
    return file_path